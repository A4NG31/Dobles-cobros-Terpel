import pandas as pd
import streamlit as st
import io
from openpyxl import load_workbook
import plotly.express as px
import plotly.graph_objects as go
from datetime import datetime, timedelta

# Configuración de página
st.set_page_config(
    page_title="GoPass - Validador Terpel", 
    page_icon="⛽", 
    layout="wide",
    initial_sidebar_state="collapsed"
)

# CSS personalizado con colores GoPass
st.markdown("""
<style>
    /* Variables de colores GoPass */
    :root {
        --gopass-blue: #1E3A8A;
        --gopass-light-blue: #3B82F6;
        --gopass-orange: #F59E0B;
        --gopass-green: #10B981;
        --gopass-red: #EF4444;
        --gopass-gray: #6B7280;
        --gopass-light-gray: #F3F4F6;
    }
    
    /* Ocultar elementos default de Streamlit */
    .stDeployButton {display: none;}
    header[data-testid="stHeader"] {display: none;}
    .stMainBlockContainer {padding-top: 1rem;}
    
    /* Header principal */
    .main-header {
        background: linear-gradient(135deg, var(--gopass-blue) 0%, var(--gopass-light-blue) 100%);
        color: white;
        padding: 2rem 0;
        margin: -1rem -1rem 2rem -1rem;
        text-align: center;
        border-radius: 0 0 20px 20px;
        box-shadow: 0 4px 20px rgba(30, 58, 138, 0.3);
    }
    
    .main-title {
        font-size: 2.5rem;
        font-weight: 700;
        margin-bottom: 0.5rem;
        text-shadow: 2px 2px 4px rgba(0,0,0,0.3);
    }
    
    .main-subtitle {
        font-size: 1.2rem;
        opacity: 0.9;
        font-weight: 300;
    }
    
    /* Cards mejoradas */
    .metric-card {
        background: white;
        padding: 1.5rem;
        border-radius: 15px;
        box-shadow: 0 4px 15px rgba(0,0,0,0.1);
        border-left: 4px solid var(--gopass-light-blue);
        margin-bottom: 1rem;
        transition: transform 0.3s ease, box-shadow 0.3s ease;
    }
    
    .metric-card:hover {
        transform: translateY(-2px);
        box-shadow: 0 8px 25px rgba(0,0,0,0.15);
    }
    
    .metric-value {
        font-size: 2rem;
        font-weight: 700;
        color: var(--gopass-blue);
        margin-bottom: 0.5rem;
    }
    
    .metric-label {
        font-size: 0.9rem;
        color: var(--gopass-gray);
        font-weight: 500;
        text-transform: uppercase;
        letter-spacing: 0.5px;
    }
    
    /* Progress bar personalizada */
    .stProgress .st-bo {
        background: linear-gradient(90deg, var(--gopass-light-blue), var(--gopass-orange));
        height: 10px;
        border-radius: 5px;
    }
    
    /* Botones mejorados */
    .stDownloadButton button, .stButton button {
        background: linear-gradient(135deg, var(--gopass-orange) 0%, #F59E0B 100%);
        color: white;
        border: none;
        border-radius: 10px;
        padding: 0.75rem 2rem;
        font-weight: 600;
        text-transform: uppercase;
        letter-spacing: 0.5px;
        transition: all 0.3s ease;
        box-shadow: 0 4px 15px rgba(245, 158, 11, 0.3);
    }
    
    .stDownloadButton button:hover, .stButton button:hover {
        transform: translateY(-2px);
        box-shadow: 0 6px 20px rgba(245, 158, 11, 0.4);
        background: linear-gradient(135deg, #D97706 0%, var(--gopass-orange) 100%);
    }
    
    /* File uploader mejorado */
    .uploadedFile {
        background: var(--gopass-light-gray);
        border: 2px dashed var(--gopass-light-blue);
        border-radius: 15px;
        padding: 2rem;
        text-align: center;
        transition: all 0.3s ease;
    }
    
    .uploadedFile:hover {
        border-color: var(--gopass-orange);
        background: rgba(59, 130, 246, 0.05);
    }
    
    /* Status badges */
    .status-normal {
        background: var(--gopass-green);
        color: white;
        padding: 0.25rem 0.75rem;
        border-radius: 20px;
        font-size: 0.8rem;
        font-weight: 600;
    }
    
    .status-doble {
        background: var(--gopass-red);
        color: white;
        padding: 0.25rem 0.75rem;
        border-radius: 20px;
        font-size: 0.8rem;
        font-weight: 600;
    }
    
    /* Alertas personalizadas */
    .custom-info {
        background: linear-gradient(135deg, rgba(59, 130, 246, 0.1), rgba(59, 130, 246, 0.05));
        border-left: 4px solid var(--gopass-light-blue);
        padding: 1rem 1.5rem;
        border-radius: 10px;
        margin: 1rem 0;
        color: var(--gopass-blue);
    }
    
    .custom-success {
        background: linear-gradient(135deg, rgba(16, 185, 129, 0.1), rgba(16, 185, 129, 0.05));
        border-left: 4px solid var(--gopass-green);
        padding: 1rem 1.5rem;
        border-radius: 10px;
        margin: 1rem 0;
        color: var(--gopass-green);
    }
    
    .custom-warning {
        background: linear-gradient(135deg, rgba(245, 158, 11, 0.1), rgba(245, 158, 11, 0.05));
        border-left: 4px solid var(--gopass-orange);
        padding: 1rem 1.5rem;
        border-radius: 10px;
        margin: 1rem 0;
        color: var(--gopass-orange);
    }
    
    /* Footer */
    .footer {
        text-align: center;
        padding: 2rem 0;
        margin-top: 3rem;
        border-top: 1px solid var(--gopass-light-gray);
        color: var(--gopass-gray);
        font-size: 0.9rem;
    }
    
    /* Dataframe styling */
    .stDataFrame {
        border-radius: 15px;
        overflow: hidden;
        box-shadow: 0 4px 15px rgba(0,0,0,0.1);
    }
    
    /* Animaciones */
    @keyframes fadeIn {
        from { opacity: 0; transform: translateY(20px); }
        to { opacity: 1; transform: translateY(0); }
    }
    
    .fade-in {
        animation: fadeIn 0.6s ease forwards;
    }
</style>
""", unsafe_allow_html=True)

# Header principal
st.markdown("""
<div class="main-header fade-in">
    <img src="https://i.imgur.com/z9xt46F.jpeg" 
         style="width: 200px; border-radius: 15px; margin-bottom: 1rem; box-shadow: 0 4px 15px rgba(0,0,0,0.3);" 
         alt="Logo GoPass">
    <div class="main-title">⛽ Validador de Dobles Cobros</div>
    <div class="main-subtitle">Gasolineras Terpel - Sistema de Detección Avanzada</div>
</div>
""", unsafe_allow_html=True)

# Inicializar session_state
if 'datos_procesados' not in st.session_state:
    st.session_state.datos_procesados = False
    st.session_state.df_final = None
    st.session_state.current_file = None
    st.session_state.stats = None

# Sección de carga de archivo
st.markdown('<div class="fade-in">', unsafe_allow_html=True)

col1, col2, col3 = st.columns([1, 2, 1])
with col2:
    st.markdown("### 📂 Cargar Base de Datos de Transacciones")
    st.markdown('<div class="custom-info">📋 Formatos soportados: Excel (.xlsx) | Tamaño máximo: 500MB</div>', unsafe_allow_html=True)
    
    uploaded_file = st.file_uploader(
        "Selecciona el archivo de transacciones Terpel",
        type=['xlsx'],
        help="Archivo Excel con las transacciones de gasolineras Terpel",
        accept_multiple_files=False
    )

st.markdown('</div>', unsafe_allow_html=True)

# Verificar cambio de archivo
if uploaded_file is not None and st.session_state.current_file != uploaded_file.name:
    st.session_state.datos_procesados = False
    st.session_state.df_final = None
    st.session_state.current_file = uploaded_file.name
    st.session_state.stats = None

# Procesamiento de datos
if uploaded_file is not None and not st.session_state.datos_procesados:
    try:
        # Verificar tamaño
        file_size_mb = len(uploaded_file.getvalue()) / (1024 * 1024)
        if file_size_mb > 500:
            st.markdown(f'<div class="custom-warning">⚠️ Archivo demasiado grande ({file_size_mb:.1f}MB). Máximo: 500MB</div>', unsafe_allow_html=True)
            st.stop()
        
        st.markdown(f'<div class="custom-success">📊 Archivo cargado exitosamente: {uploaded_file.name} ({file_size_mb:.1f}MB)</div>', unsafe_allow_html=True)
        
        # Sección de procesamiento
        st.markdown("### ⚙️ Procesando Datos")
        
        # Progress container
        progress_container = st.container()
        with progress_container:
            col1, col2, col3 = st.columns([1, 2, 1])
            with col2:
                st.markdown("**📂 Cargando archivo Excel...**")
                progress_bar = st.progress(0)
                status_text = st.empty()
        
        # Cargar archivo
        wb = load_workbook(uploaded_file, read_only=True)
        ws = wb.active
        total_filas = ws.max_row
        
        data = []
        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            data.append(row)
            if i % 1000 == 0 or i == total_filas:
                progress = int((i / total_filas) * 100)
                progress_bar.progress(progress)
                status_text.text(f"Procesando fila {i:,} de {total_filas:,}")
        
        status_text.text("✅ Archivo cargado completamente")
        
        # Crear DataFrame
        columnas = data[0]
        df = pd.DataFrame(data[1:], columns=columnas)
        
        # Guardar total bruto antes de filtros
        total_registros_bruto = len(df)
        
        # Procesamiento principal
        with col2:
            st.markdown("**🔍 Analizando dobles cobros...**")
            progress_bar = st.progress(0)
            status_text = st.empty()
        
        # Conversión de fecha y filtrado
        df['Fecha de Pago'] = pd.to_datetime(df['Fecha de Pago'], dayfirst=True, errors='coerce')
        df = df[(df['Valor Pagado'] > 0) & (df['Estado'] == 'Exitosa')].copy()
        
        # Columna de novedad
        df['Novedad'] = "NORMAL"
        
        # Ordenar por campos clave
        df.sort_values(by=['Establecimiento', 'Placa', 'Valor Pagado', 'Fecha de Pago'], inplace=True)
        df.reset_index(drop=True, inplace=True)
        
        # Detección de dobles cobros
        total = len(df)
        dobles_detectados = 0
        
        for i in range(1, total):
            # Condiciones para doble cobro
            mismo_establecimiento = df.iloc[i]['Establecimiento'] == df.iloc[i-1]['Establecimiento']
            misma_placa = df.iloc[i]['Placa'] == df.iloc[i-1]['Placa']
            mismo_valor_servicio = df.iloc[i]['Valor Servicio'] == df.iloc[i-1]['Valor Servicio']
            mismo_valor_pagado = df.iloc[i]['Valor Pagado'] == df.iloc[i-1]['Valor Pagado']
            id_diferente = df.iloc[i]['Id'] != df.iloc[i-1]['Id']
            
            # Diferencia de tiempo (10 minutos de tolerancia)
            diferencia_tiempo = abs((df.iloc[i]['Fecha de Pago'] - df.iloc[i-1]['Fecha de Pago']).total_seconds())
            dentro_tolerancia = diferencia_tiempo <= 600  # 10 minutos = 600 segundos
            
            if (mismo_establecimiento and misma_placa and mismo_valor_servicio and 
                mismo_valor_pagado and id_diferente and dentro_tolerancia):
                df.at[i, 'Novedad'] = "DOBLE COBRO"
                df.at[i-1, 'Novedad'] = "DOBLE COBRO"
                dobles_detectados += 2
            
            # Actualizar progreso
            if i % 1000 == 0 or i == total - 1:
                progress = int((i / total) * 100)
                progress_bar.progress(progress)
                status_text.text(f"Analizando registro {i:,} de {total:,} - Dobles detectados: {dobles_detectados}")
        
        status_text.text("✅ Análisis completado")
        
        # Preparar DataFrame final
        columnas_finales = ['Fecha de Pago', 'Id', 'Establecimiento', 'Placa', 'Valor Servicio', 
                           'Valor Pagado', 'Estado', 'Novedad']
        st.session_state.df_final = df[columnas_finales].copy()
        
        # Calcular estadísticas
        total_registros = len(st.session_state.df_final)
        dobles_cobros = len(st.session_state.df_final[st.session_state.df_final['Novedad'] == 'DOBLE COBRO'])
        porcentaje_dobles = (dobles_cobros/total_registros*100) if total_registros > 0 else 0
        valor_total_dobles = st.session_state.df_final[st.session_state.df_final['Novedad'] == 'DOBLE COBRO']['Valor Pagado'].sum()

        # Después de cargar y filtrar los datos
        total_registros_bruto = len(df)  # <-- los 7070 brutos
        total_registros = len(df_filtrado)  # <-- los 6268 filtrados
        
        # Al guardar estadísticas agrega el bruto
        st.session_state.stats = {
            'total_registros_bruto': total_registros_bruto,
            'total_registros': total_registros,
            'dobles_cobros': dobles_cobros,
            'porcentaje_dobles': porcentaje_dobles,
            'valor_total_dobles': valor_total_dobles,
            'registros_normales': total_registros - dobles_cobros
        }
        
        st.session_state.datos_procesados = True
        
    except Exception as e:
        st.markdown(f'<div class="custom-warning">⚠️ Error al procesar el archivo: {str(e)}</div>', unsafe_allow_html=True)
        st.stop()

# Mostrar resultados
if uploaded_file is not None and st.session_state.datos_procesados and st.session_state.df_final is not None:
    
    st.markdown("---")
    st.markdown("### 📊 Dashboard de Resultados")
    
    stats = st.session_state.stats
    
    # Métricas principales
    col1, col2, col3, col4, col5 = st.columns(5)
    
    with col1:
    st.metric(
        label="TOTAL REGISTROS BRUTOS",
        value=f"{st.session_state.stats['total_registros_bruto']:,}".replace(",", ".")
    )

    with col2:
        st.metric(
            label="TOTAL REGISTROS",
            value=f"{st.session_state.stats['total_registros']:,}".replace(",", ".")
        )
    
    with col3:
        st.metric(
            label="DOBLES COBROS",
            value=f"{st.session_state.stats['dobles_cobros']:,}".replace(",", ".")
        )
    
    with col4:
        st.metric(
            label="% DOBLES COBROS",
            value=f"{st.session_state.stats['porcentaje_dobles']:.2f}%"
        )
    
    with col5:
        st.metric(
            label="VALOR TOTAL DOBLES",
            value=f"${st.session_state.stats['valor_total_dobles']:,}".replace(",", ".")
        )
        
    # Gráficos
    col1, col2 = st.columns(2)
    
    with col1:
        # Gráfico de dona
        fig_dona = go.Figure(data=[go.Pie(
            labels=['Registros Normales', 'Dobles Cobros'],
            values=[stats['registros_normales'], stats['dobles_cobros']],
            hole=0.6,
            marker_colors=['#10B981', '#EF4444']
        )])
        fig_dona.update_layout(
            title="Distribución de Registros",
            title_font_size=16,
            title_x=0.5,
            showlegend=True,
            height=300
        )
        st.plotly_chart(fig_dona, use_container_width=True)
    
    with col2:
        # Análisis por establecimiento (top 10 con más dobles cobros)
        if stats['dobles_cobros'] > 0:
            top_establecimientos = st.session_state.df_final[
                st.session_state.df_final['Novedad'] == 'DOBLE COBRO'
            ]['Establecimiento'].value_counts().head(10)
            
            fig_bar = px.bar(
                x=top_establecimientos.values,
                y=top_establecimientos.index,
                orientation='h',
                title="Top 10 Establecimientos con Dobles Cobros",
                color=top_establecimientos.values,
                color_continuous_scale=['#F59E0B', '#EF4444']
            )
            fig_bar.update_layout(height=300, showlegend=False)
            st.plotly_chart(fig_bar, use_container_width=True)
    
    # Vista previa de datos
    st.markdown("### 📋 Vista Previa de Resultados")
    
    # Filtros
    col1, col2, col3 = st.columns(3)
    with col1:
        filtro_novedad = st.selectbox(
            "Filtrar por novedad:",
            ["Todos", "NORMAL", "DOBLE COBRO"]
        )
    
    with col2:
        num_registros = st.selectbox(
            "Registros a mostrar:",
            [50, 100, 200, 500, "Todos"]
        )
    
    # Aplicar filtros
    df_mostrar = st.session_state.df_final.copy()
    if filtro_novedad != "Todos":
        df_mostrar = df_mostrar[df_mostrar['Novedad'] == filtro_novedad]
    
    if num_registros != "Todos":
        df_mostrar = df_mostrar.head(num_registros)
    
    # Mostrar datos con estilo
    st.dataframe(
        df_mostrar,
        use_container_width=True,
        height=400
    )
    
    # Botones de acción
    col1, col2, col3 = st.columns([1, 1, 1])
    
    with col2:
        # Preparar archivo para descarga
        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            # Hoja principal con todos los datos
            st.session_state.df_final.to_excel(writer, sheet_name='Resultados_Completos', index=False)
            
            # Hoja solo con dobles cobros
            dobles_df = st.session_state.df_final[st.session_state.df_final['Novedad'] == 'DOBLE COBRO']
            dobles_df.to_excel(writer, sheet_name='Solo_Dobles_Cobros', index=False)
            
            # Hoja de resumen
            resumen_data = {
                'Métrica': ['Total Registros', 'Dobles Cobros', 'Registros Normales', '% Dobles Cobros', 'Valor Total Dobles'],
                'Valor': [stats['total_registros'], stats['dobles_cobros'], stats['registros_normales'], 
                         f"{stats['porcentaje_dobles']:.2f}%", f"${stats['valor_total_dobles']:,.0f}"]
            }
            pd.DataFrame(resumen_data).to_excel(writer, sheet_name='Resumen', index=False)
        
        buffer.seek(0)
        
        st.download_button(
            label="💾 Descargar Análisis Completo",
            data=buffer,
            file_name=f"Terpel_DoblesCobros_Analisis_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    
    # Botón para reiniciar
    if st.button("🔄 Analizar Nuevo Archivo"):
        st.session_state.datos_procesados = False
        st.session_state.df_final = None
        st.session_state.current_file = None
        st.session_state.stats = None
        st.rerun()

# Mensaje inicial
if uploaded_file is None:
    st.markdown("""
    <div class="custom-info">
        <h4>📋 Instrucciones de Uso</h4>
        <p><strong>1.</strong> Descarga la base de transacciones Terpel desde la consola GoPass</p>
        <p><strong>2.</strong> Carga el archivo Excel (.xlsx) utilizando el botón superior</p>
        <p><strong>3.</strong> El sistema detectará automáticamente los dobles cobros basado en:</p>
        <ul>
            <li>✅ Mismo Establecimiento, Placa, Valor Servicio y Valor Pagado</li>
            <li>⏰ Diferencia de tiempo ≤ 10 minutos</li>
            <li>✅ Estado = "Exitosa"</li>
            <li>🆔 ID Transaction diferentes</li>
        </ul>
        <p><strong>4.</strong> Revisa los resultados en el dashboard y descarga el análisis</p>
    </div>
    """, unsafe_allow_html=True)

# Footer
st.markdown("""
<div class="footer">
    <hr style="border: 1px solid var(--gopass-light-gray); margin: 2rem 0;">
    <p><strong>🚀 GoPass Analytics Platform</strong> | Validador de Dobles Cobros Terpel v2.0</p>
    <p>Desarrollado por <strong>Angel Torres</strong> | © 2025 GoPass</p>
</div>
""", unsafe_allow_html=True)
